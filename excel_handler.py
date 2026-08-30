from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

EXPECTED_COLUMNS = [
    "Student Name",
    "Roll Number",
    "Registration Number",
    "Amount Paid",
    "Transaction ID of your payment",
    "Payment Date",
    "Payment Time",
    "Paid To",
    "Screenshot Amount",
    "Transaction ID",
]

def read_students(path):
    path = Path(path)
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path, dtype=str)
    else:
        df = pd.read_excel(path, dtype=str)

    df.columns = [str(c).strip() for c in df.columns]

    # Accept common aliases without changing the original data.
    aliases = {
        "Student": "Student Name",
        "Name": "Student Name",
        "Expected Amount": "Amount Paid",
        "Amount": "Amount Paid",
        "UPI": "UPI ID",
        "UPI ID": "UPI ID",
    }
    df = df.rename(columns={c: aliases.get(c, c) for c in df.columns})

    if "Student Name" not in df.columns:
        raise ValueError("Student list must contain a 'Student Name' column.")
    if "Amount Paid" not in df.columns:
        raise ValueError("Student list must contain an 'Amount Paid' column.")

    for col in EXPECTED_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    return df.fillna("")


def write_template(output_path):
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(columns=EXPECTED_COLUMNS).to_excel(
        output_path,
        index=False,
        sheet_name="Student Payments",
    )

    wb = load_workbook(output_path)
    style_sheet(wb["Student Payments"])
    wb.save(output_path)
    return str(output_path)


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_cells in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(max_len + 2, 12), 42)


def style_duplicate_groups(ws):
    group_colors = [
        "DDEBF7",
        "E2F0D9",
        "FFF2CC",
        "FCE4D6",
        "E4DFEC",
        "D9EAD3",
    ]
    headers = {cell.value: cell.column for cell in ws[1]}
    transaction_id_col = headers.get("Bank Transaction ID") or headers.get(
        "Claimed Transaction ID"
    )
    status_col = headers.get("Status")

    if not transaction_id_col:
        return

    group_fills = {}
    for row in range(2, ws.max_row + 1):
        if status_col and ws.cell(row, status_col).value != "DUPLICATE":
            continue

        transaction_id = str(ws.cell(row, transaction_id_col).value or "").strip().lower()
        if not transaction_id:
            continue

        if transaction_id not in group_fills:
            color = group_colors[len(group_fills) % len(group_colors)]
            group_fills[transaction_id] = PatternFill("solid", fgColor=color)

        fill = group_fills[transaction_id]
        for cell in ws[row]:
            cell.fill = fill


def write_report(result, output_path):
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(result["verification"]).to_excel(writer, sheet_name="Payment Verification", index=False)
        pd.DataFrame(result["genuine"]).to_excel(writer, sheet_name="Genuine Payments", index=False)
        pd.DataFrame(result["invalid"]).to_excel(writer, sheet_name="Invalid Payments", index=False)
        pd.DataFrame(result["review"]).to_excel(writer, sheet_name="Needs Review", index=False)
        pd.DataFrame(result["not_found"]).to_excel(writer, sheet_name="Not Found", index=False)
        pd.DataFrame(result["unclaimed"]).to_excel(writer, sheet_name="Unclaimed Credits", index=False)
        pd.DataFrame(result["duplicates"]).to_excel(writer, sheet_name="Duplicate Conflicts", index=False)
        pd.DataFrame(list(result["summary"].items()), columns=["Metric", "Value"]).to_excel(
            writer, sheet_name="Summary", index=False
        )

    wb = load_workbook(output_path)
    for ws in wb.worksheets:
        style_sheet(ws)

    style_duplicate_groups(wb["Duplicate Conflicts"])
    style_duplicate_groups(wb["Payment Verification"])

    # Highlight status cells on the main sheet.
    ws = wb["Payment Verification"]
    headers = {c.value: c.column for c in ws[1]}
    status_col = headers.get("Status")
    if status_col:
        fills = {
            "GENUINE": PatternFill("solid", fgColor="C6EFCE"),    # Light green
            "INVALID": PatternFill("solid", fgColor="FFC7CE"),    # Light red
            "REVIEW": PatternFill("solid", fgColor="FFEB9C"),    # Light yellow
            "NOT FOUND": PatternFill("solid", fgColor="FFC000"),    # Orange
            "DUPLICATE": PatternFill("solid", fgColor="DDEBF7"),    # Light purple
        }
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, status_col)
            if cell.value in fills:
                cell.fill = fills[cell.value]

    wb.save(output_path)
    return str(output_path)
